from __future__ import annotations

import os
import tempfile
from copy import copy
from io import BytesIO
from pathlib import Path

from sqlmodel import Session
from openpyxl import Workbook, load_workbook
from openpyxl.formula.translate import Translator

from app.core.errors import BadRequestError, NotFoundError
from app.models import Customer, Sale
from app.services import sale_service

TEMPLATE_RESERVED_ITEM_ROWS = 5


def _set_cell_value(ws, r, c, val):
    cell = ws.cell(row=r, column=c)
    if type(cell).__name__ == "MergedCell":
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col, value=val)
                return
    else:
        cell.value = val


def _copy_sheet_settings(src_ws, dst_ws):
    dst_ws.sheet_format.defaultRowHeight = src_ws.sheet_format.defaultRowHeight
    dst_ws.sheet_format.defaultColWidth = src_ws.sheet_format.defaultColWidth
    dst_ws.sheet_properties = copy(src_ws.sheet_properties)
    dst_ws.page_margins = copy(src_ws.page_margins)
    dst_ws.page_setup = copy(src_ws.page_setup)
    dst_ws.print_options = copy(src_ws.print_options)
    dst_ws.freeze_panes = src_ws.freeze_panes
    dst_ws.sheet_state = src_ws.sheet_state

    try:
        dst_ws.print_title_rows = src_ws.print_title_rows
        dst_ws.print_title_cols = src_ws.print_title_cols
    except Exception:
        pass

    for col_letter, dim in src_ws.column_dimensions.items():
        dst_dim = dst_ws.column_dimensions[col_letter]
        dst_dim.width = dim.width
        dst_dim.hidden = dim.hidden
        dst_dim.bestFit = dim.bestFit
        try:
            dst_dim.outlineLevel = dim.outlineLevel
        except Exception:
            pass
        dst_dim.collapsed = dim.collapsed
        dst_dim.min = dim.min
        dst_dim.max = dim.max


def _copy_template_row(src_ws, dst_ws, src_row: int, dst_row: int, max_col: int | None = None):
    max_col = max_col or src_ws.max_column
    src_dim = src_ws.row_dimensions[src_row]
    dst_dim = dst_ws.row_dimensions[dst_row]
    dst_dim.height = src_dim.height
    dst_dim.hidden = src_dim.hidden
    try:
        dst_dim.outlineLevel = src_dim.outlineLevel
    except Exception:
        pass
    dst_dim.collapsed = src_dim.collapsed

    for c in range(1, max_col + 1):
        src = src_ws.cell(src_row, c)
        dst = dst_ws.cell(dst_row, c)

        value = src.value
        if isinstance(value, str) and value.startswith("="):
            try:
                value = Translator(value, origin=src.coordinate).translate_formula(dst.coordinate)
            except Exception:
                pass

        dst.value = value
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)

        if src.hyperlink:
            dst._hyperlink = copy(src.hyperlink)
        if src.comment:
            dst.comment = copy(src.comment)


def _copy_fixed_merges(src_ws, dst_ws, start_row: int):
    for merged_range in src_ws.merged_cells.ranges:
        if merged_range.max_row < start_row:
            dst_ws.merge_cells(str(merged_range))


def _copy_row_merges(src_ws, dst_ws, src_row: int, dst_row: int):
    for merged_range in src_ws.merged_cells.ranges:
        if merged_range.min_row == src_row and merged_range.max_row == src_row:
            dst_ws.merge_cells(
                start_row=dst_row,
                end_row=dst_row,
                start_column=merged_range.min_col,
                end_column=merged_range.max_col,
            )


def _locate_detail_columns(ws):
    start_row = 12
    col_map = {}

    for r in range(1, 30):
        for c in range(1, 30):
            cell = ws.cell(row=r, column=c)
            if type(cell).__name__ == "MergedCell":
                continue
            val = str(cell.value or "").strip()
            if val == "序号":
                col_map["index"] = c
                start_row = r + 1
            elif val in ["名称", "品名", "商品名称"]:
                col_map["name"] = c
            elif val in ["规格", "SKU"]:
                col_map["sku"] = c
            elif val == "数量":
                col_map["qty"] = c
            elif val == "单位":
                col_map["unit"] = c
            elif val == "单价":
                col_map["price"] = c
            elif val == "金额":
                col_map["amount"] = c

        if "index" in col_map and "name" in col_map:
            break

    return start_row, col_map


def _build_dynamic_sale_sheet(template_wb, *, item_count: int):
    template_ws = template_wb.active
    start_row, col_map = _locate_detail_columns(template_ws)

    reserved_rows = TEMPLATE_RESERVED_ITEM_ROWS
    separator_template_row = start_row + reserved_rows
    total_template_row = separator_template_row + 1
    footer_blank_template_row = total_template_row + 1
    footer_empty_template_row = total_template_row + 2
    signature_template_row = total_template_row + 3

    out_wb = Workbook()
    out_wb.calculation = copy(template_wb.calculation)
    out_ws = out_wb.active
    out_ws.title = template_ws.title
    _copy_sheet_settings(template_ws, out_ws)

    for r in range(1, start_row):
        _copy_template_row(template_ws, out_ws, r, r)
    _copy_fixed_merges(template_ws, out_ws, start_row)

    for idx in range(item_count):
        src_row = start_row + min(idx, reserved_rows - 1)
        dst_row = start_row + idx
        _copy_template_row(template_ws, out_ws, src_row, dst_row)
        _copy_row_merges(template_ws, out_ws, src_row, dst_row)

    separator_row = start_row + item_count
    total_row = separator_row + 1
    footer_blank_row = total_row + 1
    footer_empty_row = total_row + 2
    signature_row = total_row + 3

    for src_row, dst_row in [
        (separator_template_row, separator_row),
        (total_template_row, total_row),
        (footer_blank_template_row, footer_blank_row),
        (footer_empty_template_row, footer_empty_row),
        (signature_template_row, signature_row),
    ]:
        _copy_template_row(template_ws, out_ws, src_row, dst_row)
        _copy_row_merges(template_ws, out_ws, src_row, dst_row)

    if "amount" in col_map:
        amount_col_letter = out_ws.cell(row=start_row - 1, column=col_map["amount"]).column_letter
        out_ws[
            f"{amount_col_letter}{total_row}"] = f"=SUM({amount_col_letter}{start_row}:{amount_col_letter}{separator_row})"

    big_amount_cell = f"E{total_template_row}"
    if isinstance(template_ws[big_amount_cell].value, str) and template_ws[big_amount_cell].value.startswith("="):
        out_ws[f"E{total_row}"] = Translator(
            template_ws[big_amount_cell].value,
            origin=big_amount_cell,
        ).translate_formula(f"E{total_row}")

    return out_wb, out_ws, start_row, col_map


def _fill_sale_header(ws, *, sale, customer_phone: str, doc_type: str):
    sale_date_str = sale.sale_date.strftime("%Y-%m-%d %H:%M")

    for r in range(1, 20):
        for c in range(1, 20):
            cell = ws.cell(row=r, column=c)
            if type(cell).__name__ == "MergedCell":
                continue

            val = str(cell.value or "").strip()
            # 动态替换标题
            if "销售清单" in val:
                title = "销售清单"
                if doc_type == "quote":
                    title = "报价单"
                elif doc_type == "delivery":
                    title = "送货单"
                _set_cell_value(ws, r, c, val.replace("销售清单", title))
            elif val in ["单号：", "单号"]:
                _set_cell_value(ws, r, c + 2, sale.sale_no)
            elif val in ["日期：", "日期"]:
                _set_cell_value(ws, r, c + 2, sale_date_str)
            elif val in ["客户名称：", "客户名称", "客户"]:
                _set_cell_value(ws, r, c + 2, sale.customer_name)
            elif val in ["电话：", "电话", "联系电话：", "联系电话"]:
                _set_cell_value(ws, r, c + 2, customer_phone)

    # 如果是报价单或销售单，抹除无关紧要的物流签收字段
    if doc_type in ["quote", "sale"]:
        for r in range(12, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell_val = str(ws.cell(row=r, column=c).value or "")
                if "发货地" in cell_val or "收货人" in cell_val or "签字" in cell_val:
                    ws.cell(row=r, column=c).value = ""


def export_sale_excel(session: Session, *, sale_id: int, template_path: str | None = None, doc_type: str = "sale") -> \
tuple[bytes, str, str]:
    exists = session.get(Sale, sale_id)
    if not exists:
        raise NotFoundError("单据不存在")
    sale = sale_service.get_sale(session, sale_id)

    current_dir = Path(__file__).resolve().parent
    backend_dir = current_dir.parent.parent
    root_dir = backend_dir.parent

    possible_paths = [
        Path(template_path) if template_path else None,
        backend_dir / "打印模板.xlsx",
        root_dir / "打印模板.xlsx",
        Path("打印模板.xlsx"),
    ]

    path_to_use = None
    for p in possible_paths:
        if p and p.exists():
            path_to_use = p
            break

    if not path_to_use:
        raise BadRequestError("找不到【打印模板.xlsx】，请确认该文件已上传")

    template_wb = load_workbook(path_to_use)
    wb, ws, start_row, col_map = _build_dynamic_sale_sheet(template_wb, item_count=len(sale.items))

    customer = session.get(Customer, sale.customer_id) if getattr(sale, "customer_id", None) else None
    customer_phone = getattr(sale, "contact_phone_snapshot", None)
    if not customer_phone and customer:
        customer_phone = (
                getattr(customer, "phone", None)
                or getattr(customer, "mobile", None)
                or getattr(customer, "contact_phone", None)
        )
    customer_phone = customer_phone or "-"

    _fill_sale_header(ws, sale=sale, customer_phone=customer_phone, doc_type=doc_type)

    for idx, it in enumerate(sale.items):
        r = start_row + idx
        if "index" in col_map:
            _set_cell_value(ws, r, col_map["index"], idx + 1)
        if "name" in col_map:
            _set_cell_value(ws, r, col_map["name"], it.product_name)
        if "sku" in col_map:
            _set_cell_value(ws, r, col_map["sku"], it.sku or "")
        if "qty" in col_map:
            _set_cell_value(ws, r, col_map["qty"], float(it.qty))
        if "unit" in col_map:
            _set_cell_value(ws, r, col_map["unit"], it.unit or "")
        if "price" in col_map:
            _set_cell_value(ws, r, col_map["price"], float(it.unit_price))

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"


def export_sale_pdf(session: Session, *, sale_id: int, template_path: str | None = None,
                    doc_type: str = "sale") -> bytes:
    excel_bytes, _, _ = export_sale_excel(session, sale_id=sale_id, template_path=template_path, doc_type=doc_type)

    try:
        import pythoncom
        import win32com.client
    except ImportError:
        raise BadRequestError("请在 backend 目录下运行 pip install pywin32 以支持 PDF 转换")

    fd_xlsx, temp_xlsx = tempfile.mkstemp(suffix=".xlsx")
    with os.fdopen(fd_xlsx, "wb") as f:
        f.write(excel_bytes)

    temp_pdf = temp_xlsx.replace(".xlsx", ".pdf")

    try:
        pythoncom.CoInitialize()
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        abs_xlsx = os.path.abspath(temp_xlsx)
        abs_pdf = os.path.abspath(temp_pdf)

        wb = excel.Workbooks.Open(abs_xlsx)
        wb.ExportAsFixedFormat(0, abs_pdf)
        wb.Close(False)
        excel.Quit()
    except Exception as e:
        raise BadRequestError(f"PDF转换失败，请确认后台 Excel 进程未卡死: {str(e)}")
    finally:
        pythoncom.CoUninitialize()

    try:
        with open(temp_pdf, "rb") as f:
            pdf_bytes = f.read()
    except FileNotFoundError:
        raise BadRequestError("未能成功生成 PDF 文件")
    finally:
        if os.path.exists(temp_xlsx):
            try:
                os.remove(temp_xlsx)
            except Exception:
                pass
        if os.path.exists(temp_pdf):
            try:
                os.remove(temp_pdf)
            except Exception:
                pass

    return pdf_bytes